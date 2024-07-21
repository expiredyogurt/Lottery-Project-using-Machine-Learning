#!/usr/bin/env python
# coding: utf-8

# In[1]:


# _____         _             ____  _                  _                
#|_   _|  ___  | |_   ___    / ___|| |__    ___   ___ | | __  ___  _ __ 
#  | |   / _ \ | __| / _ \  | |    | '_ \  / _ \ / __|| |/ / / _ \| '__|
#  | |  | (_) || |_ | (_) | | |___ | | | ||  __/| (__ |   < |  __/| |   
#  |_|   \___/  \__| \___/   \____||_| |_| \___| \___||_|\_\ \___||_|   
                                                                     


# In[2]:


#this code extract the winning numbers from the website to test if its working
import requests
from bs4 import BeautifulSoup

# URL of the webpage to scrape
url = 'https://check4d.org/'

# Send a GET request to the webpage
response = requests.get(url)

# Check if the request was successful
if response.status_code != 200:
    print("Failed to retrieve webpage.")
    exit()

# Parse the HTML content of the webpage
soup = BeautifulSoup(response.text, 'html.parser')

# Find the specific div with class 'col-lg-4 col-md-4 col-sm-12 col-xs-12 col-xxs-12 lpad'
target_div = soup.find('div', class_='col-lg-4 col-md-4 col-sm-12 col-xs-12 col-xxs-12 lpad')
if target_div is None:
    print("Target div not found.")
    exit()

# Find all td elements with class 'resulttop' and 'resultbottom' within the target div
top_tds = target_div.find_all('td', class_='resulttop')
bottom_tds = target_div.find_all('td', class_='resultbottom')

# Initialize lists to store the numbers
prize_numbers = {
    '1stPrizeNo': [],
    '2ndPrizeNo': [],
    '3rdPrizeNo': [],
    'SpecialNo': [],
    'ConsolationNo': []
}

# Extract the 4-digit numbers from 'resulttop' td elements
for i, td in enumerate(top_tds):
    number = td.text.strip()
    if number.isdigit():  # Check if the text is a numeric value
        if i == 0:
            prize_numbers['1stPrizeNo'].append(number)
        elif i == 1:
            prize_numbers['2ndPrizeNo'].append(number)
        elif i == 2:
            prize_numbers['3rdPrizeNo'].append(number)
        else:
            prize_numbers['SpecialNo'].append(number)

# Extract the 4-digit numbers from 'resultbottom' td elements
for td in bottom_tds:
    number = td.text.strip()
    if number.isdigit():  # Check if the text is a numeric value
        prize_numbers['ConsolationNo'].append(number)

# Format the numbers as specified
formatted_numbers = []
for key in ['1stPrizeNo', '2ndPrizeNo', '3rdPrizeNo']:
    formatted_numbers.extend(prize_numbers[key])

for key in ['SpecialNo', 'ConsolationNo']:
    formatted_numbers.extend(prize_numbers[key])

# Print the compiled list of numbers
print("\t".join(formatted_numbers))







# In[3]:


#This final code check the website for the winning numbers then cross check it in the saved dataset to see how many winning numbers. 


# In[4]:


#This final code check the website for the winning numbers then cross check it in the saved dataset to see how many winning numbers 

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date

# Get today's date
today_date = date.today().strftime('%d %b %Y')

# URL of the webpage to scrape
url = 'https://check4d.org/'

# Send a GET request to the webpage
response = requests.get(url)

# Check if the request was successful
if response.status_code != 200:
    print("Failed to retrieve webpage.")
    exit()

# Parse the HTML content of the webpage
soup = BeautifulSoup(response.text, 'html.parser')

# Find the specific div with class 'col-lg-4 col-md-4 col-sm-12 col-xs-12 col-xxs-12 lpad'
target_div = soup.find('div', class_='col-lg-4 col-md-4 col-sm-12 col-xs-12 col-xxs-12 lpad')
if target_div is None:
    print("Target div not found.")
    exit()

# Find all td elements with class 'resulttop' and 'resultbottom' within the target div
top_tds = target_div.find_all('td', class_='resulttop')
bottom_tds = target_div.find_all('td', class_='resultbottom')

# Function to extract only the 4-digit numbers from the text and convert to int
def extract_4_digit_number(text):
    # Extract digits from the text
    digits = ''.join(filter(str.isdigit, text))
    # Check if the number has exactly 4 digits
    if len(digits) == 4:
        return int(digits)  # Convert to integer
    else:
        return None

# Initialize dictionary to store the numbers
prize_numbers = {
    '1stPrizeNo': extract_4_digit_number(top_tds[0].text.strip()),
    '2ndPrizeNo': extract_4_digit_number(top_tds[1].text.strip()),
    '3rdPrizeNo': extract_4_digit_number(top_tds[2].text.strip()),
    'SpecialNo1': extract_4_digit_number(bottom_tds[0].text.strip()),
    'SpecialNo2': extract_4_digit_number(bottom_tds[1].text.strip()),
    'SpecialNo3': extract_4_digit_number(bottom_tds[2].text.strip()),
    'SpecialNo4': extract_4_digit_number(bottom_tds[3].text.strip()),
    'SpecialNo5': extract_4_digit_number(bottom_tds[4].text.strip()),
    'SpecialNo6': extract_4_digit_number(bottom_tds[5].text.strip()),
    'SpecialNo7': extract_4_digit_number(bottom_tds[6].text.strip()),
    'SpecialNo8': extract_4_digit_number(bottom_tds[7].text.strip()),
    'SpecialNo9': extract_4_digit_number(bottom_tds[8].text.strip()),
    'SpecialNo10': extract_4_digit_number(bottom_tds[9].text.strip()),
    'ConsolationNo1': extract_4_digit_number(bottom_tds[10].text.strip()),
    'ConsolationNo2': extract_4_digit_number(bottom_tds[11].text.strip()),
    'ConsolationNo3': extract_4_digit_number(bottom_tds[12].text.strip()),
    'ConsolationNo4': extract_4_digit_number(bottom_tds[13].text.strip()),
    'ConsolationNo5': extract_4_digit_number(bottom_tds[14].text.strip()),
    'ConsolationNo6': extract_4_digit_number(bottom_tds[15].text.strip()),
    'ConsolationNo7': extract_4_digit_number(bottom_tds[16].text.strip()),
    'ConsolationNo8': extract_4_digit_number(bottom_tds[17].text.strip()),
    'ConsolationNo9': extract_4_digit_number(bottom_tds[18].text.strip()),
    'ConsolationNo10': extract_4_digit_number(bottom_tds[19].text.strip()),
}

import pandas as pd
from datetime import date

# Get today's date
today_date = date.today().strftime('%d %b %Y')

# Define the file paths for the CSV files
deepL_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/4D DeepL Prediction {today_date}.csv'
xgb_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/4D XGB Prediction {today_date}.csv'
nn_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/4D NN Prediction {today_date}.csv'
xgb_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/4D Random Forest Reg Prediction {today_date}.csv'
lotteryAI_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/LotteryAI Prediction {today_date}.csv'

# Read and preprocess the CSV files and ensure numbers are stripped of whitespace and converted to int
def read_and_preprocess_csv(csv_file_path):
    df = pd.read_csv(csv_file_path, dtype=str)  # Read CSV with string dtype to preserve leading zeros
    # Strip whitespace from all columns and convert to int, skipping rows with non-numeric data
    df = df.apply(lambda col: col.map(lambda x: int(x.strip()) if isinstance(x, str) and x.strip().isdigit() else None))
    # Drop rows with any NaN values (rows with non-numeric data)
    df = df.dropna()
    return df

# Read and preprocess the DeepL CSV file
deepL_df = read_and_preprocess_csv(deepL_csv_file_path)

# Read and preprocess the XGB CSV file
xgb_df = read_and_preprocess_csv(xgb_csv_file_path)

# Read and preprocess the NN V3 CSV file
nn_v3_df = read_and_preprocess_csv(nn_v3_csv_file_path)

# Read and preprocess the XGB V3 CSV file
xgb_v3_df = read_and_preprocess_csv(xgb_v3_csv_file_path)

# Read and preprocess the LotteryAI CSV file
lotteryAI_df = read_and_preprocess_csv(lotteryAI_csv_file_path)

# Remove any None values from the dictionary
prize_numbers = {key: value for key, value in prize_numbers.items() if value is not None}

# Print the extracted numbers with labels
print("Numbers obtained from the website:")
for label, number in prize_numbers.items():
    print(f"{label}: {number}")

import pandas as pd

# Function to check for matches between the extracted numbers and the numbers in a DataFrame
def check_matches(df, csv_name):
    matches_found = False
    for index, row in df.iterrows():
        for col, number in row.items():  # Using items() instead of iteritems()
            if isinstance(number, int) and number in prize_numbers.values():
                matches_found = True
                print(f"Match found in {csv_name} at row {index + 1}: {col} {number}")
    if not matches_found:
        print(f"No matches found in {csv_name}.")




# Check for matches in each CSV file
check_matches(deepL_df, "DeepL CSV")
check_matches(xgb_df, "XGB CSV")
check_matches(nn_v3_df, "NN Predicition CSV")
check_matches(xgb_v3_df, "Random Forest Regression CSV")
check_matches(lotteryAI_df, "LotteryAI CSV")


# In[5]:


#This final code check the website for the winning numbers then cross check it in the saved dataset to see how many winning numbers 

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date

# Get today's date
today_date = date.today().strftime('%d %b %Y')

# URL of the webpage to scrape
url = 'https://check4d.org/'

# Send a GET request to the webpage
response = requests.get(url)

# Check if the request was successful
if response.status_code != 200:
    print("Failed to retrieve webpage.")
    exit()

# Parse the HTML content of the webpage
soup = BeautifulSoup(response.text, 'html.parser')

# Find the specific div with class 'col-lg-4 col-md-4 col-sm-12 col-xs-12 col-xxs-12 lpad'
target_div = soup.find('div', class_='col-lg-4 col-md-4 col-sm-12 col-xs-12 col-xxs-12')
if target_div is None:
    print("Target div not found.")
    exit()

# Find all td elements with class 'resulttop' and 'resultbottom' within the target div
top_tds = target_div.find_all('td', class_='resulttop')
bottom_tds = target_div.find_all('td', class_='resultbottom')

# Function to extract only the 4-digit numbers from the text and convert to int
def extract_4_digit_number(text):
    # Extract digits from the text
    digits = ''.join(filter(str.isdigit, text))
    # Check if the number has exactly 4 digits
    if len(digits) == 4:
        return int(digits)  # Convert to integer
    else:
        return None

# Initialize dictionary to store the numbers
prize_numbers = {
    '1stPrizeNo': extract_4_digit_number(top_tds[0].text.strip()),
    '2ndPrizeNo': extract_4_digit_number(top_tds[1].text.strip()),
    '3rdPrizeNo': extract_4_digit_number(top_tds[2].text.strip()),
    'SpecialNo1': extract_4_digit_number(bottom_tds[0].text.strip()),
    'SpecialNo2': extract_4_digit_number(bottom_tds[1].text.strip()),
    'SpecialNo3': extract_4_digit_number(bottom_tds[2].text.strip()),
    'SpecialNo4': extract_4_digit_number(bottom_tds[3].text.strip()),
    'SpecialNo5': extract_4_digit_number(bottom_tds[4].text.strip()),
    'SpecialNo6': extract_4_digit_number(bottom_tds[5].text.strip()),
    'SpecialNo7': extract_4_digit_number(bottom_tds[6].text.strip()),
    'SpecialNo8': extract_4_digit_number(bottom_tds[7].text.strip()),
    'SpecialNo9': extract_4_digit_number(bottom_tds[8].text.strip()),
    'SpecialNo10': extract_4_digit_number(bottom_tds[9].text.strip()),
    'ConsolationNo1': extract_4_digit_number(bottom_tds[10].text.strip()),
    'ConsolationNo2': extract_4_digit_number(bottom_tds[11].text.strip()),
    'ConsolationNo3': extract_4_digit_number(bottom_tds[12].text.strip()),
    'ConsolationNo4': extract_4_digit_number(bottom_tds[13].text.strip()),
    'ConsolationNo5': extract_4_digit_number(bottom_tds[14].text.strip()),
    'ConsolationNo6': extract_4_digit_number(bottom_tds[15].text.strip()),
    'ConsolationNo7': extract_4_digit_number(bottom_tds[16].text.strip()),
    'ConsolationNo8': extract_4_digit_number(bottom_tds[17].text.strip()),
    'ConsolationNo9': extract_4_digit_number(bottom_tds[18].text.strip()),
    'ConsolationNo10': extract_4_digit_number(bottom_tds[19].text.strip()),
}

import pandas as pd
from datetime import date

# Get today's date
today_date = date.today().strftime('%d %b %Y')

# Define the file paths for the CSV files
deepL_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC DeepL Prediction {today_date}.csv'
xgb_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC XGB Prediction {today_date}.csv'
nn_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC NN Prediction {today_date}.csv'
xgb_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC Random Forest Reg Prediction {today_date}.csv'
lotteryAI_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC LotteryAI Prediction {today_date}.csv'

# Read and preprocess the CSV files and ensure numbers are stripped of whitespace and converted to int
def read_and_preprocess_csv(csv_file_path):
    df = pd.read_csv(csv_file_path, dtype=str)  # Read CSV with string dtype to preserve leading zeros
    # Strip whitespace from all columns and convert to int, skipping rows with non-numeric data
    df = df.apply(lambda col: col.map(lambda x: int(x.strip()) if isinstance(x, str) and x.strip().isdigit() else None))
    # Drop rows with any NaN values (rows with non-numeric data)
    df = df.dropna()
    return df

# Read and preprocess the DeepL CSV file
deepL_df = read_and_preprocess_csv(deepL_csv_file_path)

# Read and preprocess the XGB CSV file
xgb_df = read_and_preprocess_csv(xgb_csv_file_path)

# Read and preprocess the NN V3 CSV file
nn_v3_df = read_and_preprocess_csv(nn_v3_csv_file_path)

# Read and preprocess the XGB V3 CSV file
xgb_v3_df = read_and_preprocess_csv(xgb_v3_csv_file_path)

# Read and preprocess the LotteryAI CSV file
lotteryAI_df = read_and_preprocess_csv(lotteryAI_csv_file_path)

# Remove any None values from the dictionary
prize_numbers = {key: value for key, value in prize_numbers.items() if value is not None}

# Print the extracted numbers with labels
print("Numbers obtained from the website:")
for label, number in prize_numbers.items():
    print(f"{label}: {number}")

import pandas as pd

# Function to check for matches between the extracted numbers and the numbers in a DataFrame
def check_matches(df, csv_name):
    matches_found = False
    for index, row in df.iterrows():
        for col, number in row.items():  # Using items() instead of iteritems()
            if isinstance(number, int) and number in prize_numbers.values():
                matches_found = True
                print(f"Match found in {csv_name} at row {index + 1}: {col} {number}")
    if not matches_found:
        print(f"No matches found in {csv_name}.")




# Check for matches in each CSV file
check_matches(deepL_df, "DeepL CSV")
check_matches(xgb_df, "XGB CSV")
check_matches(nn_v3_df, "NN Predicition CSV")
check_matches(xgb_v3_df, "Random Forest Regression CSV")
check_matches(lotteryAI_df, "LotteryAI CSV")


# In[6]:


#This code checks the permutation of the winning number against the CSV files


# In[7]:


import itertools

def check_permutations(prize_numbers, csv_name, df):
    # Set to store unique matched permutations
    matched_permutations = set()
    
    # Iterate over each (label, number) pair in the prize_numbers dictionary
    for label, num in prize_numbers.items():
        # Generate permutations of the digits in the number
        num_permutations = itertools.permutations(str(num))
        
        # Iterate over each permutation
        for perm in num_permutations:
            # Convert the permutation back to an integer
            perm_num = int(''.join(perm))
            
            # Check if the permutation exists in the CSV file
            if perm_num in df.values:
                # Add the matched permutation to the set
                matched_permutations.add((label, perm_num))
    
    # Print the matched permutations
    if matched_permutations:
        print(f"Matched permutations in {csv_name}:")
        for label, perm in matched_permutations:
            print(f"{label}: {perm}")
    else:
        print(f"No matches found in {csv_name}.")

# Example usage:
check_permutations(prize_numbers, "DeepL CSV", deepL_df)
check_permutations(prize_numbers, "XGB CSV", xgb_df)
check_permutations(prize_numbers, "NN Prediction CSV", nn_v3_df)
check_permutations(prize_numbers, "Random Forest Regression CSV", xgb_v3_df)
check_permutations(prize_numbers, "LotteryAI CSV", lotteryAI_df)


# In[8]:


#highlight the cells with the winning number 


# In[9]:


import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import itertools
from datetime import date

# Define the function to read and preprocess the CSV files
def read_and_preprocess_csv(csv_file_path):
    df = pd.read_csv(csv_file_path)  # Read CSV file
    # Preprocess the DataFrame if needed
    return df

def check_permutations(prize_numbers, csv_name, df, output_directory):
    # Set to store unique matched permutations
    matched_permutations = set()
    
    # Iterate over each (label, number) pair in the prize_numbers dictionary
    for label, num in prize_numbers.items():
        # Generate permutations of the digits in the number
        num_permutations = itertools.permutations(str(num))
        
        # Iterate over each permutation
        for perm in num_permutations:
            # Convert the permutation back to an integer
            perm_num = int(''.join(perm))
            
            # Check if the permutation exists in the CSV file
            if perm_num in df.values:
                # Add the matched permutation to the set
                matched_permutations.add((label, perm_num))
    
    # Print the matched permutations
    if matched_permutations:
        print(f"Matched permutations in {csv_name}:")
        for label, perm in matched_permutations:
            print(f"{label}: {perm}")
            
        # Convert the CSV file to an XLSX file and highlight the matched cells
        highlight_matched_cells(csv_name, df, matched_permutations, output_directory)
    else:
        print(f"No matches found in {csv_name}.")

def highlight_matched_cells(csv_name, df, matched_permutations, output_directory):
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write DataFrame column labels to Excel file
    ws.append(list(df.columns))
    
    # Write DataFrame to Excel file
    for r in df.values.tolist():
        ws.append(r)
    
    # Highlight cells with yellow fill color for matched permutations
    for label, perm_num in matched_permutations:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                if value == perm_num:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Save the workbook in the specified output directory
    output_filename = os.path.join(output_directory, f"{os.path.splitext(os.path.basename(csv_name))[0]}.xlsx")
    wb.save(output_filename)
    print(f"Highlighted XLSX file saved: {output_filename}")

# Define today's date with format d%b%y%
today_date = date.today().strftime('%d %b %Y')

# Define the archive directory path
archive_directory = '/Users/expiredyogurt/Desktop/Lottery Project/Archive/'

# Define the file paths for the CSV files
deepL_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/4D DeepL Prediction {today_date}.csv'
xgb_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/4D XGB Prediction {today_date}.csv'
nn_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/4D NN Prediction {today_date}.csv'
lotteryAI_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/LotteryAI Prediction {today_date}.csv'
random_forest_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/4D Random Forest Reg Prediction {today_date}.csv'

# Read and preprocess the CSV files
deepL_df = read_and_preprocess_csv(deepL_csv_file_path)
xgb_df = read_and_preprocess_csv(xgb_csv_file_path)
nn_v3_df = read_and_preprocess_csv(nn_v3_csv_file_path)
lotteryAI_df = read_and_preprocess_csv(lotteryAI_csv_file_path)
random_forest_df = read_and_preprocess_csv(random_forest_file_path)

# Example usage:
check_permutations(prize_numbers, deepL_csv_file_path, deepL_df, archive_directory)
check_permutations(prize_numbers, xgb_csv_file_path, xgb_df, archive_directory)
check_permutations(prize_numbers, nn_v3_csv_file_path, nn_v3_df, archive_directory)
check_permutations(prize_numbers, lotteryAI_csv_file_path, lotteryAI_df, archive_directory)
check_permutations(prize_numbers, random_forest_file_path, random_forest_df, archive_directory)


# In[10]:


# ____          __  __          ____         _    ____  _                  _                
#|  _ \   __ _ |  \/  |  __ _  / ___|  __ _ (_)  / ___|| |__    ___   ___ | | __  ___  _ __ 
#| | | | / _` || |\/| | / _` || |     / _` || | | |    | '_ \  / _ \ / __|| |/ / / _ \| '__|
#| |_| || (_| || |  | || (_| || |___ | (_| || | | |___ | | | ||  __/| (__ |   < |  __/| |   
#|____/  \__,_||_|  |_| \__,_| \____| \__,_||_|  \____||_| |_| \___| \___||_|\_\ \___||_| 


# In[ ]:





# In[11]:


#This final code check the website for the winning numbers then cross check it in the saved dataset to see how many winning numbers 

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date

# Function to extract only the 4-digit numbers from the text and convert to int
def extract_4_digit_number(text):
    # Extract digits from the text
    digits = ''.join(filter(str.isdigit, text))
    # Check if the number has exactly 4 digits
    if len(digits) == 4:
        return int(digits)  # Convert to integer
    else:
        return None

# Get today's date
today_date = date.today().strftime('%d %b %Y')

# URL of the webpage to scrape
url = 'https://check4d.org/'

# Send a GET request to the webpage
response = requests.get(url)

# Check if the request was successful
if response.status_code != 200:
    print("Failed to retrieve webpage.")
    exit()

# Parse the HTML content of the webpage
soup = BeautifulSoup(response.text, 'html.parser')

# Initialize dictionary to store the numbers
prize_numbers = {
    '1stPrizeNo': extract_4_digit_number(soup.find('td', id='dp1').text.strip()),
    '2ndPrizeNo': extract_4_digit_number(soup.find('td', id='dp2').text.strip()),
    '3rdPrizeNo': extract_4_digit_number(soup.find('td', id='dp3').text.strip()),
    'SpecialNo1': extract_4_digit_number(soup.find('td', id='ds1').text.strip()),
    'SpecialNo2': extract_4_digit_number(soup.find('td', id='ds2').text.strip()),
    'SpecialNo3': extract_4_digit_number(soup.find('td', id='ds3').text.strip()),
    'SpecialNo4': extract_4_digit_number(soup.find('td', id='ds4').text.strip()),
    'SpecialNo5': extract_4_digit_number(soup.find('td', id='ds5').text.strip()),
    'SpecialNo6': extract_4_digit_number(soup.find('td', id='ds6').text.strip()),
    'SpecialNo7': extract_4_digit_number(soup.find('td', id='ds7').text.strip()),
    'SpecialNo8': extract_4_digit_number(soup.find('td', id='ds8').text.strip()),
    'SpecialNo9': extract_4_digit_number(soup.find('td', id='ds9').text.strip()),
    'SpecialNo10': extract_4_digit_number(soup.find('td', id='ds10').text.strip()),
    'ConsolationNo1': extract_4_digit_number(soup.find('td', id='dc1').text.strip()),
    'ConsolationNo2': extract_4_digit_number(soup.find('td', id='dc2').text.strip()),
    'ConsolationNo3': extract_4_digit_number(soup.find('td', id='dc3').text.strip()),
    'ConsolationNo4': extract_4_digit_number(soup.find('td', id='dc4').text.strip()),
    'ConsolationNo5': extract_4_digit_number(soup.find('td', id='dc5').text.strip()),
    'ConsolationNo6': extract_4_digit_number(soup.find('td', id='dc6').text.strip()),
    'ConsolationNo7': extract_4_digit_number(soup.find('td', id='dc7').text.strip()),
    'ConsolationNo8': extract_4_digit_number(soup.find('td', id='dc8').text.strip()),
    'ConsolationNo9': extract_4_digit_number(soup.find('td', id='dc9').text.strip()),
    'ConsolationNo10': extract_4_digit_number(soup.find('td', id='dc10').text.strip()),
}



import pandas as pd
from datetime import date

# Get today's date
today_date = date.today().strftime('%d %b %Y')

# Define the file paths for the CSV files
deepL_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC DeepL Prediction {today_date}.csv'
xgb_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC XGB Prediction {today_date}.csv'
nn_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC NN Prediction {today_date}.csv'
xgb_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC Random Forest Reg Prediction {today_date}.csv'
lotteryAI_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC LotteryAI Prediction {today_date}.csv'
lstm_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC LSTM Top 3 Prediction {today_date}.csv'


# Read and preprocess the CSV files and ensure numbers are stripped of whitespace and converted to int
def read_and_preprocess_csv(csv_file_path):
    df = pd.read_csv(csv_file_path, dtype=str)  # Read CSV with string dtype to preserve leading zeros
    # Strip whitespace from all columns and convert to int, skipping rows with non-numeric data
    df = df.apply(lambda col: col.map(lambda x: int(x.strip()) if isinstance(x, str) and x.strip().isdigit() else None))
    # Drop rows with any NaN values (rows with non-numeric data)
    df = df.dropna()
    return df

# Read and preprocess the DeepL CSV file
deepL_df = read_and_preprocess_csv(deepL_csv_file_path)

# Read and preprocess the XGB CSV file
xgb_df = read_and_preprocess_csv(xgb_csv_file_path)

# Read and preprocess the NN V3 CSV file
nn_v3_df = read_and_preprocess_csv(nn_v3_csv_file_path)

# Read and preprocess the XGB V3 CSV file
xgb_v3_df = read_and_preprocess_csv(xgb_v3_csv_file_path)

# Read and preprocess the LotteryAI CSV file
lotteryAI_df = read_and_preprocess_csv(lotteryAI_csv_file_path)

# Read and preprocess the LSTM CSV file
lstm_df = read_and_preprocess_csv(lstm_csv_file_path)



# Remove any None values from the dictionary
prize_numbers = {key: value for key, value in prize_numbers.items() if value is not None}

# Print the extracted numbers with labels
print("Numbers obtained from the website:")
for label, number in prize_numbers.items():
    print(f"{label}: {number}")

import pandas as pd

# Function to check for matches between the extracted numbers and the numbers in a DataFrame
def check_matches(df, csv_name):
    matches_found = False
    for index, row in df.iterrows():
        for col, number in row.items():  # Using items() instead of iteritems()
            if isinstance(number, int) and number in prize_numbers.values():
                matches_found = True
                print(f"Match found in {csv_name} at row {index + 1}: {col} {number}")
    if not matches_found:
        print(f"No matches found in {csv_name}.")




# Check for matches in each CSV file
check_matches(deepL_df, "DeepL CSV")
check_matches(xgb_df, "XGB CSV")
check_matches(nn_v3_df, "NN Predicition CSV")
check_matches(xgb_v3_df, "Random Forest Regression CSV")
check_matches(lotteryAI_df, "LotteryAI CSV")
check_matches(lstm_df, "LSTM Top 3 CSV")



# In[12]:


#This checks the winning numbers permutation with the csv file


# In[13]:


import itertools

def check_permutations(prize_numbers, csv_name, df):
    # Set to store unique matched permutations
    matched_permutations = set()
    
    # Iterate over each (label, number) pair in the prize_numbers dictionary
    for label, num in prize_numbers.items():
        # Generate permutations of the digits in the number
        num_permutations = itertools.permutations(str(num))
        
        # Iterate over each permutation
        for perm in num_permutations:
            # Convert the permutation back to an integer
            perm_num = int(''.join(perm))
            
            # Check if the permutation exists in the CSV file
            if perm_num in df.values:
                # Add the matched permutation to the set
                matched_permutations.add((label, perm_num))
    
    # Print the matched permutations
    if matched_permutations:
        print(f"Matched permutations in {csv_name}:")
        for label, perm in matched_permutations:
            print(f"{label}: {perm}")
    else:
        print(f"No matches found in {csv_name}.")

# Example usage:
check_permutations(prize_numbers, "DeepL CSV", deepL_df)
check_permutations(prize_numbers, "XGB CSV", xgb_df)
check_permutations(prize_numbers, "NN Prediction CSV", nn_v3_df)
check_permutations(prize_numbers, "Random Forest Regression CSV", xgb_v3_df)
check_permutations(prize_numbers, "LotteryAI CSV", lotteryAI_df)
check_permutations(prize_numbers, "LSTM CSV", lstm_df)



# In[14]:


#get the winning number and append it to the csv file automatically 


# In[15]:


import pandas as pd
from datetime import date
import datetime


# Get today's date
today_date = date.today()

# Assuming today_date_yyyymmdd_int is a datetime object
today_date_yyyymmdd_int = datetime.datetime.now().strftime('%Y-%m-%d')

# Define the file path for the existing CSV file
existing_csv_file_path = '/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC Latest Result.csv'

# Read the existing CSV file
existing_df = pd.read_csv(existing_csv_file_path)

# Find the index of the new row
new_row_index = len(existing_df)

# Fill the first and second columns with '1' and today's date in the new row
existing_df.loc[new_row_index, existing_df.columns[0]] = 1

# Assuming today_date_yyyymmdd_int is a datetime object
today_date_yyyymmdd_int = datetime.datetime.now().strftime('%Y-%m-%d')

existing_df.loc[new_row_index, existing_df.columns[1]] = today_date_yyyymmdd_int

# Fill the prize numbers from the dictionary list starting from the 3rd column onwards
# Adjusting the index for the column where we start adding the prize numbers
start_column_index = 2
for i, num in enumerate(prize_numbers.values()):
    existing_df.loc[new_row_index, existing_df.columns[start_column_index + i]] = num

# Fill the last column with 'Yes' in the new row
existing_df.loc[new_row_index, existing_df.columns[-1]] = 'Yes'

# Drop duplicates
existing_df = existing_df.drop_duplicates()

# Write the updated DataFrame back to the CSV file
existing_df.to_csv(existing_csv_file_path, index=False)

print("Winning numbers updated in the original CSV file:", existing_csv_file_path)


# In[16]:


#highlight the cells with the winning number 


# In[17]:


import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import itertools
from datetime import date

# Define the function to read and preprocess the CSV files
def read_and_preprocess_csv(csv_file_path):
    df = pd.read_csv(csv_file_path)  # Read CSV file
    # Preprocess the DataFrame if needed
    return df

def check_permutations(prize_numbers, csv_name, df, output_directory):
    # Set to store unique matched permutations
    matched_permutations = set()
    
    # Iterate over each (label, number) pair in the prize_numbers dictionary
    for label, num in prize_numbers.items():
        # Generate permutations of the digits in the number
        num_permutations = itertools.permutations(str(num))
        
        # Iterate over each permutation
        for perm in num_permutations:
            # Convert the permutation back to an integer
            perm_num = int(''.join(perm))
            
            # Check if the permutation exists in the CSV file
            if perm_num in df.values:
                # Add the matched permutation to the set
                matched_permutations.add((label, perm_num))
    
    # Print the matched permutations
    if matched_permutations:
        print(f"Matched permutations in {csv_name}:")
        for label, perm in matched_permutations:
            print(f"{label}: {perm}")
            
        # Convert the CSV file to an XLSX file and highlight the matched cells
        highlight_matched_cells(csv_name, df, matched_permutations, output_directory)
    else:
        print(f"No matches found in {csv_name}.")

def highlight_matched_cells(csv_name, df, matched_permutations, output_directory):
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write DataFrame column labels to Excel file
    ws.append(list(df.columns))
    
    # Write DataFrame to Excel file
    for r in df.values.tolist():
        ws.append(r)
    
    # Highlight cells with yellow fill color for matched permutations
    for label, perm_num in matched_permutations:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                if value == perm_num:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Save the workbook in the specified output directory
    output_filename = os.path.join(output_directory, f"{os.path.splitext(os.path.basename(csv_name))[0]}.xlsx")
    wb.save(output_filename)
    print(f"Highlighted XLSX file saved: {output_filename}")

# Define today's date with format d%b%y%
today_date = date.today().strftime('%d %b %Y')

# Define the archive directory path
archive_directory = '/Users/expiredyogurt/Desktop/Lottery Project/DMC/Archive/'

# Define the file paths for the CSV files
deepL_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC DeepL Prediction {today_date}.csv'
xgb_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC XGB Prediction {today_date}.csv'
nn_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC NN Prediction {today_date}.csv'
lotteryAI_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC LotteryAI Prediction {today_date}.csv'
random_forest_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC Random Forest Reg Prediction {today_date}.csv'
lstm_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/DMC/DMC LSTM Top 3 Prediction {today_date}.csv'


# Read and preprocess the CSV files
deepL_df = read_and_preprocess_csv(deepL_csv_file_path)
xgb_df = read_and_preprocess_csv(xgb_csv_file_path)
nn_v3_df = read_and_preprocess_csv(nn_v3_csv_file_path)
lotteryAI_df = read_and_preprocess_csv(lotteryAI_csv_file_path)
random_forest_df = read_and_preprocess_csv(random_forest_file_path)
lstm_df = read_and_preprocess_csv(lstm_csv_file_path)


# Example usage:
check_permutations(prize_numbers, deepL_csv_file_path, deepL_df, archive_directory)
check_permutations(prize_numbers, xgb_csv_file_path, xgb_df, archive_directory)
check_permutations(prize_numbers, nn_v3_csv_file_path, nn_v3_df, archive_directory)
check_permutations(prize_numbers, lotteryAI_csv_file_path, lotteryAI_df, archive_directory)
check_permutations(prize_numbers, random_forest_file_path, random_forest_df, archive_directory)
check_permutations(prize_numbers, lstm_csv_file_path, lstm_df, archive_directory)



# In[18]:


# __  __                                        
#|  \/  |  __ _   __ _  _ __   _   _  _ __ ___  
#| |\/| | / _` | / _` || '_ \ | | | || '_ ` _ \ 
#| |  | || (_| || (_| || | | || |_| || | | | | |
#|_|  |_| \__,_| \__, ||_| |_| \__,_||_| |_| |_|
#                |___/                          


# In[ ]:





# In[19]:


#This final code check the website for the winning numbers then cross check it in the saved dataset to see how many winning numbers 

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import date

# Function to extract only the 4-digit numbers from the text and convert to int
def extract_4_digit_number(text):
    # Extract digits from the text
    digits = ''.join(filter(str.isdigit, text))
    # Check if the number has exactly 4 digits
    if len(digits) == 4:
        return int(digits)  # Convert to integer
    else:
        return None

# Get today's date
today_date = date.today().strftime('%d %b %Y')

# URL of the webpage to scrape
url = 'https://check4d.org/'

# Send a GET request to the webpage
response = requests.get(url)

# Check if the request was successful
if response.status_code != 200:
    print("Failed to retrieve webpage.")
    exit()

# Parse the HTML content of the webpage
soup = BeautifulSoup(response.text, 'html.parser')

# Find the specific div with class 'col-lg-4 col-md-4 col-sm-12 col-xs-12 col-xxs-12 rpad'
target_div = soup.find('div', class_='col-lg-4 col-md-4 col-sm-12 col-xs-12 col-xxs-12 rpad')

# Find all td elements containing numbers within the target div
all_tds = target_div.find_all('td', class_='resulttop') + target_div.find_all('td', class_='resultbottom')

# Extract only 4-digit numbers from the text
extracted_numbers = [extract_4_digit_number(td.text.strip()) for td in all_tds]

# Remove 'None' values
extracted_numbers = [num for num in extracted_numbers if num is not None]

# Initialize dictionary to store the numbers
Magnum_prize_numbers = {
    '1stPrizeNo': None,
    '2ndPrizeNo': None,
    '3rdPrizeNo': None,
    'SpecialNo1': None,
    'SpecialNo2': None,
    'SpecialNo3': None,
    'SpecialNo4': None,
    'SpecialNo5': None,
    'SpecialNo6': None,
    'SpecialNo7': None,
    'SpecialNo8': None,
    'SpecialNo9': None,
    'SpecialNo10': None,
    'ConsolationNo1': None,
    'ConsolationNo2': None,
    'ConsolationNo3': None,
    'ConsolationNo4': None,
    'ConsolationNo5': None,
    'ConsolationNo6': None,
    'ConsolationNo7': None,
    'ConsolationNo8': None,
    'ConsolationNo9': None,
    'ConsolationNo10': None
}

# Update the dictionary with the extracted numbers
for label, number in zip(Magnum_prize_numbers.keys(), extracted_numbers):
    Magnum_prize_numbers[label] = number

# Convert the dictionary to a DataFrame
df = pd.DataFrame(Magnum_prize_numbers.items(), columns=['Prize', 'WinningNumber'])
import pandas as pd
from datetime import date

# Get today's date
today_date = date.today().strftime('%d %b %Y')

# Define the file paths for the CSV files
deepL_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum DeepL Prediction {today_date}.csv'
xgb_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum XGB Prediction {today_date}.csv'
nn_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum NN Prediction {today_date}.csv'
xgb_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum Random Forest Reg Prediction {today_date}.csv'
lotteryAI_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum LotteryAI Prediction {today_date}.csv'

# Read and preprocess the CSV files and ensure numbers are stripped of whitespace and converted to int
def read_and_preprocess_csv(csv_file_path):
    df = pd.read_csv(csv_file_path, dtype=str)  # Read CSV with string dtype to preserve leading zeros
    # Strip whitespace from all columns and convert to int, skipping rows with non-numeric data
    df = df.apply(lambda col: col.map(lambda x: int(x.strip()) if isinstance(x, str) and x.strip().isdigit() else None))
    # Drop rows with any NaN values (rows with non-numeric data)
    df = df.dropna()
    return df

# Read and preprocess the DeepL CSV file
deepL_df = read_and_preprocess_csv(deepL_csv_file_path)

# Read and preprocess the XGB CSV file
xgb_df = read_and_preprocess_csv(xgb_csv_file_path)

# Read and preprocess the NN V3 CSV file
nn_v3_df = read_and_preprocess_csv(nn_v3_csv_file_path)

# Read and preprocess the XGB V3 CSV file
xgb_v3_df = read_and_preprocess_csv(xgb_v3_csv_file_path)

# Read and preprocess the LotteryAI CSV file
lotteryAI_df = read_and_preprocess_csv(lotteryAI_csv_file_path)

# Remove any None values from the dictionary
prize_numbers = {key: value for key, value in prize_numbers.items() if value is not None}

# Print the extracted numbers with labels
print("Numbers obtained from the website:")
for label, number in Magnum_prize_numbers.items():
    print(f"{label}: {number}")

import pandas as pd

# Function to check for matches between the extracted numbers and the numbers in a DataFrame
def check_matches(df, csv_name):
    matches_found = False
    for index, row in df.iterrows():
        for col, number in row.items():  # Using items() instead of iteritems()
            if isinstance(number, int) and number in Magnum_prize_numbers.values():
                matches_found = True
                print(f"Match found in {csv_name} at row {index + 1}: {col} {number}")
    if not matches_found:
        print(f"No matches found in {csv_name}.")




# Check for matches in each CSV file
check_matches(deepL_df, "DeepL CSV")
check_matches(xgb_df, "XGB CSV")
check_matches(nn_v3_df, "NN Predicition CSV")
check_matches(xgb_v3_df, "Random Forest Regression CSV")
check_matches(lotteryAI_df, "LotteryAI CSV")


# In[ ]:





# In[20]:


#this code checks the permutation of the winning number against the CSV files


# In[21]:


import itertools

def check_permutations(prize_numbers, csv_name, df):
    # Set to store unique matched permutations
    matched_permutations = set()
    
    # Iterate over each (label, number) pair in the prize_numbers dictionary
    for label, num in prize_numbers.items():
        # Generate permutations of the digits in the number
        num_permutations = itertools.permutations(str(num))
        
        # Iterate over each permutation
        for perm in num_permutations:
            # Convert the permutation back to an integer
            perm_num = int(''.join(perm))
            
            # Check if the permutation exists in the DataFrame
            if perm_num in df.values:
                # Add the matched permutation to the set
                matched_permutations.add((label, perm_num))
    
    # Print the matched permutations
    if matched_permutations:
        print(f"Matched permutations in {csv_name}:")
        for label, perm in matched_permutations:
            print(f"{label}: {perm}")
    else:
        print(f"No matches found in {csv_name}.")

# Example usage:
check_permutations(Magnum_prize_numbers, "DeepL CSV", deepL_df)
check_permutations(Magnum_prize_numbers, "XGB CSV", xgb_df)
check_permutations(Magnum_prize_numbers, "NN Prediction CSV", nn_v3_df)
check_permutations(Magnum_prize_numbers, "Random Forest Regression CSV", xgb_v3_df)
check_permutations(Magnum_prize_numbers, "LotteryAI CSV", lotteryAI_df)


# In[22]:


#Save the latest result into the CSV file


# In[23]:


import pandas as pd
from datetime import date

# Get today's date
today_date = date.today()

# Convert today's date to 'yyyymmdd' format as an integer
today_date_yyyymmdd_int = int(today_date.strftime('%Y%m%d'))

# Define the file path for the existing CSV file
existing_csv_file_path = '/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum Latest Result.csv'

# Read the existing CSV file
existing_df = pd.read_csv(existing_csv_file_path)

# Find the index of the new row
new_row_index = len(existing_df)

# Fill the first and second columns with '1' and today's date in the new row
existing_df.loc[new_row_index, existing_df.columns[0]] = 1
existing_df.loc[new_row_index, existing_df.columns[1]] = today_date_yyyymmdd_int

# Fill the prize numbers from the dictionary list starting from the 3rd column onwards
# Adjusting the index for the column where we start adding the prize numbers
start_column_index = 2
for i, num in enumerate(Magnum_prize_numbers.values()):
    existing_df.loc[new_row_index, existing_df.columns[start_column_index + i]] = num

# Fill the last column with 'Yes' in the new row
existing_df.loc[new_row_index, existing_df.columns[-1]] = 'Yes'

# Drop duplicates
existing_df = existing_df.drop_duplicates()

# Write the updated DataFrame back to the CSV file
existing_df.to_csv(existing_csv_file_path, index=False)

print("Winning numbers updated in the original CSV file:", existing_csv_file_path)


# In[24]:


#highlight the winning numbers in the csv file 


# In[25]:


import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import itertools
from datetime import date

# Define the function to read and preprocess the CSV files
def read_and_preprocess_csv(csv_file_path):
    df = pd.read_csv(csv_file_path)  # Read CSV file
    # Preprocess the DataFrame if needed
    return df

def check_permutations(Magnum_prize_numbers, csv_name, df, output_directory):
    # Set to store unique matched permutations
    matched_permutations = set()
    
    # Iterate over each (label, number) pair in the prize_numbers dictionary
    for label, num in Magnum_prize_numbers.items():
        # Generate permutations of the digits in the number
        num_permutations = itertools.permutations(str(num))
        
        # Iterate over each permutation
        for perm in num_permutations:
            # Convert the permutation back to an integer
            perm_num = int(''.join(perm))
            
            # Check if the permutation exists in the CSV file
            if perm_num in df.values:
                # Add the matched permutation to the set
                matched_permutations.add((label, perm_num))
    
    # Print the matched permutations
    if matched_permutations:
        print(f"Matched permutations in {csv_name}:")
        for label, perm in matched_permutations:
            print(f"{label}: {perm}")
            
        # Convert the CSV file to an XLSX file and highlight the matched cells
        highlight_matched_cells(csv_name, df, matched_permutations, output_directory)
    else:
        print(f"No matches found in {csv_name}.")

def highlight_matched_cells(csv_name, df, matched_permutations, output_directory):
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write DataFrame column labels to Excel file
    ws.append(list(df.columns))
    
    # Write DataFrame to Excel file
    for r in df.values.tolist():
        ws.append(r)
    
    # Highlight cells with yellow fill color for matched permutations
    for label, perm_num in matched_permutations:
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                if value == perm_num:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Save the workbook in the specified output directory
    output_filename = os.path.join(output_directory, f"{os.path.splitext(os.path.basename(csv_name))[0]}.xlsx")
    wb.save(output_filename)
    print(f"Highlighted XLSX file saved: {output_filename}")

# Define today's date with format d%b%y%
today_date = date.today().strftime('%d %b %Y')

# Define the archive directory path
archive_directory = '/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Archive/'

# Define the file paths for the CSV files
deepL_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum DeepL Prediction {today_date}.csv'
xgb_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum XGB Prediction {today_date}.csv'
nn_v3_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum NN Prediction {today_date}.csv'
lotteryAI_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum LotteryAI Prediction {today_date}.csv'
lstm_csv_file_path = f'/Users/expiredyogurt/Desktop/Lottery Project/Magnum/Magnum LSTM Top 3 Prediction {today_date}.csv'


# Read and preprocess the CSV files
deepL_df = read_and_preprocess_csv(deepL_csv_file_path)
xgb_df = read_and_preprocess_csv(xgb_csv_file_path)
nn_v3_df = read_and_preprocess_csv(nn_v3_csv_file_path)
lotteryAI_df = read_and_preprocess_csv(lotteryAI_csv_file_path)
lstm_df = read_and_preprocess_csv(lstm_csv_file_path)


# Example usage:
check_permutations(prize_numbers, deepL_csv_file_path, deepL_df, archive_directory)
check_permutations(prize_numbers, xgb_csv_file_path, xgb_df, archive_directory)
check_permutations(prize_numbers, nn_v3_csv_file_path, nn_v3_df, archive_directory)
check_permutations(prize_numbers, lotteryAI_csv_file_path, lotteryAI_df, archive_directory)
check_permutations(prize_numbers, lstm_csv_file_path, lstm_df, archive_directory)



# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




